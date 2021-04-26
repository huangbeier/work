# -*- coding: utf-8 -*-
#   __author__:黄贝尔
#   2021-04-24
from openpyxl import load_workbook

from config.VarConfig import excelpath


class Excel_tools:

    def read_work_book(self,filename):
        '''
        加载文件
        :param filename: excel文件地址
        '''
        try:
            self.workbook=load_workbook(filename)
        except Exception as e:
            print(e)

    def read_work_sheet(self,sheetname):
        '''
        通过表名获取表
        :param sheetname: 表名
        :return:
        '''
        try:
            self.sheet=self.workbook[sheetname]
        except Exception as e:
            print(e)

    def get_max_rows(self):
        '''
        获取行数，从1开始
        :return:
        '''
        return self.sheet.max_row

    def get_max_col(self):
        '''
        获取列数，从1开始
        :return:
        '''
        return self.sheet.max_column

    def get_row_data(self,row):
        '''
        获取某一行的数据
        :param row: 行数
        :return:
        '''
        col = self.sheet.max_column
        row_data = []
        for i in range(1, col + 1):
            cell_value = self.sheet.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    def get_specific_data(self,row,col):
        '''
        获取具体单元格的值
        :return:
        '''
        cell_value=self.sheet.cell(row=row,column=col).value
        return cell_value

    def get_all_data(self):
        '''
        获取sheet中所有内容
        :return:
        '''
        row1 = self.sheet.max_row
        row_datas = []
        for i in range(1, row1 + 1):
            value = self.get_row_data(i)
            row_datas.append(value)
        return row_datas

    def get_all_dic_data(self):
        '''
        将第一行和下面每一条数据拼接成字典，每一行被一个字典包括，大字典的key是行数
        :return:
        '''
        frist = self.get_row_data(1)
        row1 = self.sheet.max_row
        row_datas= {}
        for i in range(2, row1 + 1):
            value = self.get_row_data(i)
            if value[0]!=None:
                datas=dict(zip(frist,value))
                row_datas[i]=datas
        return row_datas

    def write_specific_data(self,row,col,value):
        self.sheet.cell(row=row,column=col,value=value)
        self.workbook.save(excelpath)
        self.workbook.close()


if __name__ == '__main__':
    data=Excel_tools()
    data.read_work_book(excelpath)
    data.read_work_sheet('Sheet1')
    # print(data.get_all_data())
    print(data.get_all_dic_data())
    #print(data.get_row_data(3))
    # data.write_specific_data(5,5,111)
    # print(data.get_specific_data(5,5))



